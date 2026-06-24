import os
import io
import json
import shutil
from datetime import datetime
import pytz
import jdatetime
import openpyxl
import requests
import uvicorn
from fastapi.responses import FileResponse
from difflib import SequenceMatcher


from fastapi import FastAPI, Depends, HTTPException, Request, File, UploadFile
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import List
from apscheduler.schedulers.background import BackgroundScheduler
from contextlib import asynccontextmanager

import database as db

# 🎯 دیکشنری تبدیل نام کاربری به نام محترمانه
USER_MAPPING = {
    "Hadi": "آقا هادی لطفی",
    "AmirAKS9": "امیر آقا ",
    "Nima": "آقا نیما",
    "Naser": "آقا ناصر",
    "gemany": "آقا ساجد",
    "Sana": "آقا سعید",
    "Hamid": "آقا حمید",
    "alisaj": "علی آقا سجادی",
    "Alims": "علی آقا متولیان",
    "مسعود": "آقا مسعود",
    "ایران_رویایی": "آقا نادر",
    "Hadisajadi": "آقا هادی متولیان",
    "Amir_Rainbow": "امیر آقا عباسی"
}

def get_persian_name(username):
    if not username: return "کاربر ناشناس"
    # این خط کاری می‌کند که کوچکی یا بزرگی حروف تاثیری در پیدا کردن نام نداشته باشد
    mapping_lower = {k.lower(): v for k, v in USER_MAPPING.items()}
    return mapping_lower.get(username.lower().strip(), username)

def to_persian_num(num_str):
    mapping = str.maketrans('0123456789', '۰۱۲۳۴۵۶۷۸۹')
    return str(num_str).translate(mapping)


# --- تنظیمات اتصال به بله ---
BALE_TOKEN = "928514616:u3lR097wIz127f4g4W0GXRyN9KJT5kADmlI"
BALE_CHAT_ID = "@Golchine_Akhbar"
ADMIN_BALE_ID = "189389617"

# --- تنظیمات API-Sports ---
API_SPORTS_KEY = "91a19f0ef86a021c15fb02f28539fe86"


# دیکشنری ترجمه نام تیم‌ها (برای تطبیق دیتابیس فارسی با نتایج انگلیسی)
TEAM_NAME_MAPPING = {
    "ایران": "Iran",
    "ایران": "Iran", 
    "نیوزیلند": "New Zealand", 
    "فرانسه": "France", 
    "سنگال": "Senegal", "عربستان": "Saudi Arabia", "اروگوئه": "Uruguay", "مکزیک": "Mexico", "کره جنوبی": "South Korea", "کانادا": "Canada", "قطر": "Qatar", "آمریکا": "USA", "هائیتی": "Haiti", "برزیل": "Brazil", "استرالیا": "Australia", "آلمان": "Germany", "هلند": "Netherlands", "ساحل عاج": "Ivory Coast", "سوئد": "Sweden", "اسپانیا": "Spain", "اسپانیای": "Spain", "بلژیک": "Belgium", "عراق": "Iraq", "آرژانتین": "Argentina", "اتریش": "Austria", "پرتغال": "Portugal", "انگلیس": "England", "غنا": "Ghana", "ازبکستان": "Uzbekistan", "جمهوری چک": "Czech Republic", "سوئیس": "Switzerland", "اسکاتلند": "Scotland", "ترکیه": "Turkey", "اکوادور": "Ecuador", "تونس": "Tunisia", "نروژ": "Norway", "اردن": "Jordan", "پاناما": "Panama", "کلمبیا": "Colombia", "آفریقای جنوبی": "South Africa", "بوسنی و هرزگوین": "Bosnia and Herzegovina", "مراکش": "Morocco", "پاراگوئه": "Paraguay", "کوراسائو": "Curacao", "ژاپن": "Japan", "مصر": "Egypt", "کیپ ورد": "Cape Verde", "الجزایر": "Algeria", "جمهوری کنگو": "Congo", "کرواسی": "Croatia" } 
def is_team_match(db_team_en, api_team_name):
    """تابع کمکی برای تشخیص شباهت دو نام حتی در صورت وجود پسوند و پیشوند"""
    if not db_team_en or not api_team_name: 
        return False
    
    db_team = db_team_en.lower().strip()
    api_team = str(api_team_name).lower().strip()
    
    # حالت اول: نام یکی دقیقاً در دل دیگری باشد (مثل iran در iran (islamic republic))
    if db_team in api_team or api_team in db_team:
        return True
        
    # حالت دوم: شباهت املایی بالای ۷۵ درصد باشد
    ratio = SequenceMatcher(None, db_team, api_team).ratio()
    return ratio >= 0.75

def fetch_and_update_from_api(db_session: Session, target_date_str: str):
    """
    دریافت نتایج از API-Football و آپدیت خودکار دیتابیس
    فرمت ورودی: YYYY-MM-DD
    """
    url = f"https://v3.football.api-sports.io/fixtures?date={target_date_str}"
    headers = {'x-apisports-key': API_SPORTS_KEY}
    
    try:
        response = requests.get(url, headers=headers, timeout=60)
        
        # بررسی خطاهای مربوط به خود API (مثل اتمام اعتبار رایگان)
        if response.status_code != 200:
            return f"❌ خطای سرور API. کد خطا: {response.status_code}"
            
        data = response.json()
        
        # بررسی محدودیت درخواست‌های روزانه
        if data.get('errors') and 'requests' in data['errors'].get('rateLimit', ''):
            return "❌ محدودیت ۱۰۰ درخواست رایگان امروز شما در API-Football به پایان رسیده است."
            
        if not data.get('response'):
            return "⚠️ هیچ مسابقه‌ای در این تاریخ در سرور جهانی یافت نشد."
            
        api_fixtures = data['response']
        updated_count = 0
        
        # فقط بازی‌های پیش‌رو که در دیتابیس هستند را چک می‌کنیم
        pending_matches = db_session.query(db.Match).filter(db.Match.status == "upcoming").all()
        
        for match in pending_matches:
            # تبدیل نام فارسی به انگلیسی از روی دیکشنری
            home_en = TEAM_NAME_MAPPING.get(match.home_team.strip())
            away_en = TEAM_NAME_MAPPING.get(match.away_team.strip())
            
            if not home_en or not away_en:
                continue 
                
            # جستجو در نتایج API
            for api_match in api_fixtures:
                api_home = api_match['teams']['home']['name']
                api_away = api_match['teams']['away']['name']
                
                # استفاده از تابع هوشمند برای مقایسه نام تیم‌ها
                if is_team_match(home_en, api_home) and is_team_match(away_en, api_away):
                    status_short = api_match['fixture']['status']['short']
                    
                    # کدهای پایان بازی در API-Football: 
                    # FT (تمام وقت)، AET (پایان وقت اضافه)، PEN (پایان پنالتی)
                    if status_short in ['FT', 'PEN', 'AET']:
                        match.actual_home_goals = api_match['goals']['home']
                        match.actual_away_goals = api_match['goals']['away']
                        match.status = "finished"
                        updated_count += 1
                        
                        db_session.commit()
                        calculate_leaderboard_data(db_session)
                        
                        # ارسال کارنامه کاربران به بله
                        try:
                            msg_text = generate_bale_summary_message(db_session, match.id)
                            if msg_text: 
                                send_bale_notification(msg_text)
                        except Exception as e:
                            print(f"Bale API error: {e}")
                            
                    break # وقتی بازی پیدا شد، حلقه جستجو برای این بازی را بشکن
                    
        return f"✅ سینک با موفقیت انجام شد. {updated_count} مسابقه ثبت و جدول آپدیت گردید."
        
    except requests.exceptions.Timeout:
        return "❌ خطای Timeout: سرور خارجی در زمان مناسب پاسخ نداد."
    except Exception as e:
        return f"❌ خطای غیرمنتظره در ارتباط با API: {str(e)}"

# 🌟 تابع ارسال پیام (همراه با سیستم دیباگ و پردازش دکمه‌ها)
def send_bale_notification(message_text: str, target_chat_id=None, reply_markup=None):
    chat = target_chat_id if target_chat_id else BALE_CHAT_ID
    url = f"https://tapi.bale.ai/bot{BALE_TOKEN}/sendMessage"
    payload = {"chat_id": chat, "text": message_text}
   
    if reply_markup:
        payload["reply_markup"] = json.dumps(reply_markup)
       
    try:
        response = requests.post(url, json=payload, timeout=10)
        if response.status_code != 200:
            print(f"❌ Bale API Error: {response.text}")
    except Exception as e:
        print(f"❌ Connection Error: {e}")

# 🌟 تابع ساخت و ارسال منوی اصلی کاربری
def send_user_main_menu(chat_id: str):
    menu_buttons = {
        "inline_keyboard": [
            [
                {"text": "⚽️ مسابقات پیش‌رو", "callback_data": "user_upcoming_matches"},
                {"text": "🏆 جدول رده‌بندی لیگ", "callback_data": "user_leaderboard"}
            ],
            [
                {"text": "🕵️‍♂️ اتاق شفاف‌سازی (۳ بازی بعدی)", "callback_data": "user_rivals_preds"},
                {"text": "👥 کارنامه رقبا (همه بازی‌ها)", "callback_data": "user_list_public"}
            ],
            [
                {"text": "📜 قوانین و امتیازدهی", "callback_data": "user_rules"}
            ]
        ]
    }
    welcome_text = "سلام! به ربات دستیار لیگ پیش‌بینی خوش آمدید. 👋\n\n👇 لطفاً یکی از گزینه‌های زیر را انتخاب کنید:"
    send_bale_notification(welcome_text, target_chat_id=chat_id, reply_markup=menu_buttons)

# --- سیستم بک‌آپ و زمان‌بندی ---
BACKUP_DIR = "data/backups"
if not os.path.exists(BACKUP_DIR): 
    os.makedirs(BACKUP_DIR)

def backup_database():
    try:
        now_str = jdatetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        shutil.copy2("data/football.db", f"{BACKUP_DIR}/football_backup_{now_str}.db")
    except Exception: 
        pass

iran_nz_warning_sent = False
def check_iran_nz_match():
    global iran_nz_warning_sent
    if iran_nz_warning_sent: 
        return
        
    db_session = db.SessionLocal()
    try:
        match = db_session.query(db.Match).filter(
            db.Match.home_team.contains('ایران'), 
            db.Match.away_team.contains('نیوزلند')
        ).first()
        
        if match and match.timestamp and match.status == 'upcoming':
            current_ts = datetime.now(pytz.timezone("Asia/Tehran")).timestamp()
            time_left = match.timestamp - current_ts
            if 0 < time_left <= 5 * 3600:
                send_bale_notification("🚨 **یادآوری مهم مسابقات!**\n\nتنها **۵ ساعت** تا شروع مسابقه حساس **ایران ⚡️ نیوزلند** باقی مانده است!\n⏳ فرم ثبت نتایج ۱۵ دقیقه قبل از سوت آغاز قفل خواهد شد.")
                iran_nz_warning_sent = True
    finally:
        db_session.close()

prompted_matches = set()
def check_finished_matches_prompt():
    db_session = db.SessionLocal()
    try:
        current_ts = datetime.now(pytz.timezone("Asia/Tehran")).timestamp()
        pending_matches = db_session.query(db.Match).filter(
            db.Match.status == "upcoming", 
            db.Match.timestamp != None, 
            db.Match.timestamp + 7200 <= current_ts
        ).all()
        
        for m in pending_matches:
            if m.id not in prompted_matches:
                msg = f"🔔 **مدیر عزیز، زمان مسابقه زیر به پایان رسیده است:**\n⚔️ **{m.home_team} - {m.away_team}**\n\n" \
                      f"الگوهای کلیک‌شدنی برای ثبت نتیجه سریع:\n" \
                      f"🔹 مساوی صفر - صفر: `/set_{m.id}_0_0`\n" \
                      f"🔹 برد یک - صفر میزبان: `/set_{m.id}_1_0`\n" \
                      f"🔹 برد دو - یک میزبان: `/set_{m.id}_2_1`\n" \
                      f"🔹 برد صفر - یک میهمان: `/set_{m.id}_0_1`\n\n" \
                      f"الگوی دستی: `/set_{m.id}_[میزبان]_[میهمان]`"
                send_bale_notification(msg, target_chat_id=ADMIN_BALE_ID)
                prompted_matches.add(m.id)
    finally:
        db_session.close()

scheduler = BackgroundScheduler()
scheduler.add_job(backup_database, 'cron', hour=3, minute=0)
scheduler.add_job(check_iran_nz_match, 'interval', minutes=5)
scheduler.add_job(check_finished_matches_prompt, 'interval', minutes=5)
scheduler.start()

# خط تعریف اپلیکیشن
app = FastAPI(title="سیستم پیش‌بینی فوتبال")

# 🌟 این خط احتمالاً پاک شده است، حتماً باید زیر app باشد:
templates = Jinja2Templates(directory="templates")

# کدهای استارت‌آپ
@app.on_event("startup")
def startup_db_check():
    db.Base.metadata.create_all(bind=db.engine)
    with db.engine.begin() as conn:
        try: 
            conn.execute(text("ALTER TABLE users ADD COLUMN previous_rank INTEGER DEFAULT 1;"))
        except Exception: 
            pass
            
        try:
            conn.execute(text("ALTER TABLE users ADD COLUMN username TEXT;"))
            conn.execute(text("UPDATE users SET username = name WHERE username IS NULL;"))
        except Exception: 
            pass


        # 🌟 تغییر جدید: اضافه کردن ستون زمان به جدول پیش‌بینی‌ها بدون حذف اطلاعات قبلی
        try:
            conn.execute(text("ALTER TABLE predictions ADD COLUMN submit_time TEXT DEFAULT 'نامشخص';"))
        except Exception:
            pass


# کدهای PWA که قبلا اضافه کردیم
@app.get("/manifest.json", include_in_schema=False)
def get_manifest():
    return FileResponse("static/manifest.json", media_type="application/json")

@app.get("/sw.js", include_in_schema=False)
def get_service_worker():
    return FileResponse("static/sw.js", media_type="application/javascript")

@app.get("/service-worker.js", include_in_schema=False) 
def get_service_worker_alt(): 
    return FileResponse("static/sw.js", media_type="application/javascript") 

def get_db():
    db_session = db.SessionLocal()
    try: 
        yield db_session
    finally: 
        db_session.close()

def get_tehran_timestamp(j_date_str, time_str):
    try:
        clean_date = str(j_date_str).split(' ')[0].replace('-', '/')
        y, m, d = map(int, clean_date.split('/'))
        time_parts = str(time_str).split(':')
        dt_jalali = jdatetime.datetime(y, m, d, int(time_parts[0]), int(time_parts[1]), 0)
        tehran_tz = pytz.timezone("Asia/Tehran")
        return tehran_tz.localize(dt_jalali.togregorian()).timestamp()
    except Exception: 
        return None

def log_action(db_session: Session, request: Request, user_name: str, action: str, details: str):
    ip = request.client.host if request.client else "Unknown"
    ua = request.headers.get("user-agent", "Unknown")
    now_str = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M:%S")
    db_session.add(db.AuditLog(user_name=user_name, action=action, details=details, ip_address=ip, user_agent=ua, timestamp=now_str))
    db_session.commit()

class BulkDeleteRequest(BaseModel): 
    match_ids: List[int]

class MatchResultItem(BaseModel): 
    match_id: int
    actual_home: int
    actual_away: int

class BulkFinishRequest(BaseModel): 
    results: List[MatchResultItem]

def calculate_leaderboard_data(db_session):
    users = db_session.query(db.User).all()
    
    # 🌟 تنظیم مبدا جدید محاسبات: ۲۶ خرداد ساعت ۰۴:۰۰ صبح به وقت تهران
    try:
        current_year = jdatetime.datetime.now().year
        # ماه ۳ (خرداد)، روز ۲۶، ساعت ۴، دقیقه ۰، ثانیه ۰
        dt_jalali = jdatetime.datetime(current_year, 3, 26, 4, 0, 0)
        tehran_tz = pytz.timezone("Asia/Tehran")
        threshold_ts = tehran_tz.localize(dt_jalali.togregorian()).timestamp()
    except Exception:
        threshold_ts = 0 # در صورت بروز خطای پیش‌بینی نشده، کل بازی‌ها محاسبه شوند

    # دریافت تمام بازی‌های تمام‌شده‌ای که زمان آن‌ها بعد از ۴ صبح ۲۶ خرداد است
    finished_matches = db_session.query(db.Match).filter(
        db.Match.status == "finished", 
        db.Match.timestamp >= threshold_ts
    ).all()
    
    prize_setting = db_session.query(db.SystemSetting).filter(db.SystemSetting.key == "total_prize").first()
    total_prize = float(prize_setting.value) if prize_setting else 0.0

    leaderboard_data = []
    for u in users:
        stats = {"exact": 0, "diff": 0, "winner": 0, "wrong": 0, "missed": 0, "score": 0}
        preds = {p.match_id: p for p in db_session.query(db.Prediction).filter(db.Prediction.user_id == u.id).all()}
        
        for fm in finished_matches:
            if fm.id not in preds: 
                stats["missed"] += 1
            else:
                p = preds[fm.id]
                a_diff = fm.actual_home_goals - fm.actual_away_goals
                p_diff = p.predicted_home_goals - p.predicted_away_goals
                
                if p.predicted_home_goals == fm.actual_home_goals and p.predicted_away_goals == fm.actual_away_goals:
                    stats["exact"] += 1
                    stats["score"] += 3
                elif p_diff == a_diff:
                    stats["diff"] += 1
                    stats["score"] += 2
                elif (a_diff > 0 and p_diff > 0) or (a_diff < 0 and p_diff < 0):
                    stats["winner"] += 1
                    stats["score"] += 1
                else: 
                    stats["wrong"] += 1
                    
        u.score = stats["score"]
        prev_rank = getattr(u, 'previous_rank', 1) or 1
        leaderboard_data.append({"id": u.id, "name": u.name, "username": u.username, "score": stats["score"], "previous_rank": prev_rank, "prize": 0.0, "trend": "-", **stats})
       
    db_session.commit()
    leaderboard_data.sort(key=lambda x: x["score"], reverse=True)
   
    current_rank = 1
    for i, row in enumerate(leaderboard_data):
        if i > 0 and leaderboard_data[i]["score"] < leaderboard_data[i-1]["score"]: 
            current_rank = i + 1
        row["rank"] = current_rank
        
        if row["rank"] < row["previous_rank"]: 
            row["trend"] = "up"
        elif row["rank"] > row["previous_rank"]: 
            row["trend"] = "down"
        else: 
            row["trend"] = "stable"

    if total_prize > 0 and leaderboard_data:
        from collections import defaultdict
        rank_groups = defaultdict(list)
        for row in leaderboard_data: 
            rank_groups[row["rank"]].append(row)
            
        first_place_users = rank_groups[1]
        distinct_ranks = sorted(rank_groups.keys())
        second_place_users = rank_groups[distinct_ranks[1]] if len(distinct_ranks) > 1 else []

        if len(first_place_users) >= 2:
            share = total_prize / len(first_place_users)
            for u in first_place_users: 
                u["prize"] = round(share, 2)
        elif len(first_place_users) == 1:
            u1 = first_place_users[0]
            if len(second_place_users) == 1:
                u2 = second_place_users[0]
                total_pts = u1["score"] + u2["score"]
                if total_pts > 0:
                    u1["prize"] = round((u1["score"] / total_pts) * total_prize, 2)
                    u2["prize"] = round((u2["score"] / total_pts) * total_prize, 2)
                else: 
                    u1["prize"] = u2["prize"] = round(total_prize / 2, 2)
            elif len(second_place_users) >= 2:
                u1["prize"] = round(0.55 * total_prize, 2)
                share_45 = (0.45 * total_prize) / len(second_place_users)
                for u in second_place_users: 
                    u["prize"] = round(share_45, 2)
            else: 
                u1["prize"] = total_prize

    return leaderboard_data

def generate_bale_summary_message(db_session: Session, finished_match_id: int) -> str:
    match = db_session.query(db.Match).filter(db.Match.id == finished_match_id).first()
    if not match: return ""
       
    lb_data = calculate_leaderboard_data(db_session)
    user_stats = {item['id']: item for item in lb_data}

    ah_fa = to_persian_num(match.actual_home_goals)
    aa_fa = to_persian_num(match.actual_away_goals)
    
    msg_parts = [f"🏁 سوت پایان! {match.home_team} {ah_fa}-{aa_fa} {match.away_team}\n"]
    
    predictions = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == finished_match_id).all()
    predictions.sort(key=lambda x: user_stats.get(x[1].id, {}).get('score', 0), reverse=True)
    
    medals = ["🥇", "🥈", "🥉"]
    for idx, (pred, user) in enumerate(predictions):
        medal = medals[idx] if idx < 3 else "👤"
        stats = user_stats.get(user.id, {})
        
        ph = pred.predicted_home_goals
        pa = pred.predicted_away_goals
        ah = match.actual_home_goals
        aa = match.actual_away_goals
        
        if ph == ah and pa == aa: pt_txt = "۳ امتیاز کامل"
        elif (ph - pa) == (ah - aa): pt_txt = "۲ امتیاز"
        elif (ph > pa and ah > aa) or (ph < pa and ah < aa): pt_txt = "۱ امتیاز"
        else: pt_txt = "بدون امتیاز"

        dn = get_persian_name(user.username if user.username else user.name)
        total_score_fa = to_persian_num(stats.get('score', 0))
        ph_fa, pa_fa = to_persian_num(ph), to_persian_num(pa)
        
        msg_parts.append(f"{medal} {dn} ({total_score_fa} امتیاز) | حدس: {ph_fa}-{pa_fa} ({pt_txt})")

    msg_parts.append("➖➖➖")
    
    next_match = db_session.query(db.Match).filter(db.Match.status == "upcoming").order_by(db.Match.timestamp.asc()).first()
    if next_match:
        time_left_str = "نامشخص"
        if next_match.timestamp:
            ts_secs = int((datetime.fromtimestamp(next_match.timestamp) - datetime.now()).total_seconds())
            if ts_secs > 0:
                h, r = divmod(ts_secs, 3600)
                m = r // 60
                time_left_str = f"{to_persian_num(h)} ساعت و {to_persian_num(m)} دقیقه" if h > 0 else f"{to_persian_num(m)} دقیقه"
            else: time_left_str = "زمان تمام شده!"

        msg_parts.append(f" بعدی: {next_match.home_team} 🆚 {next_match.away_team}")
        msg_parts.append(f"⏳ زمان : {time_left_str}")
        msg_parts.append("👀 پیش‌بینی‌ها:")
        
        next_preds = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == next_match.id).all()
        for npred, nuser in next_preds:
            dn = get_persian_name(nuser.username if nuser.username else nuser.name)
            
            nph = npred.predicted_home_goals
            npa = npred.predicted_away_goals
            
            # تشخیص تیمی که کاربر پیش‌بینی کرده می‌برد
            if nph > npa:
                pred_winner = next_match.home_team
            elif npa > nph:
                pred_winner = next_match.away_team
            else:
                pred_winner = "مساوی"
                
            msg_parts.append(f"👤 {dn}: ({to_persian_num(nph)}-{to_persian_num(npa)}) {pred_winner}")
    return "\n".join(msg_parts)

# --- مسیرهای مربوط به FastAPI ---

@app.get("/")
def home(request: Request): 
    return templates.TemplateResponse(request=request, name="index.html", context={"request": request})

def generate_live_bale_message(db_session: Session, match_id: int, live_home: int, live_away: int) -> str:
    match = db_session.query(db.Match).filter(db.Match.id == match_id).first()
    if not match: return ""

    # دریافت جدول پایه (بدون در نظر گرفتن این بازی)
    base_lb = calculate_leaderboard_data(db_session)
    base_stats = {item['id']: item for item in base_lb}

    preds = db_session.query(db.Prediction, db.User).join(db.User).filter(db.Prediction.match_id == match_id).all()
    temp_users = []

    for pred, user in preds:
        ph, pa = pred.predicted_home_goals, pred.predicted_away_goals
        pt = 0
        if ph == live_home and pa == live_away: pt = 3
        elif (ph - pa) == (live_home - live_away): pt = 2
        elif (ph > pa and live_home > live_away) or (ph < pa and live_home < live_away): pt = 1

        current_score = base_stats.get(user.id, {}).get('score', 0)
        temp_users.append({
            "name": get_persian_name(user.username if user.username else user.name),
            "score": current_score + pt,
            "ph_fa": to_persian_num(ph),
            "pa_fa": to_persian_num(pa),
            "pt_txt": "۳ امتیاز کامل" if pt==3 else "۲ امتیاز" if pt==2 else "۱ امتیاز" if pt==1 else "بدون امتیاز"
        })

    # مرتب‌سازی کاربران بر اساس امتیاز لحظه‌ای
    temp_users.sort(key=lambda x: x['score'], reverse=True)

    msg_parts = [
        f"تغییر جدول با نتیجه فعلی ⚽️",
        f"نتیجه فعلی {to_persian_num(live_home)}-{to_persian_num(live_away)}\n"
    ]

    for idx, tu in enumerate(temp_users):
        medal = ["🥇", "🥈", "🥉"][idx] if idx < 3 else "👤"
        msg_parts.append(f"{medal} {tu['name']} ({to_persian_num(tu['score'])} امتیاز) | حدس: {tu['ph_fa']}-{tu['pa_fa']} ({tu['pt_txt']})")

    return "\n".join(msg_parts)

class LiveUpdateReq(BaseModel):
    match_id: int
    home_goals: int
    away_goals: int

@app.post("/matches/live-update")
def live_update_match(req: LiveUpdateReq, db_session: Session = Depends(get_db)):
    match = db_session.query(db.Match).filter(db.Match.id == req.match_id).first()
    if not match: return {"status": "error"}

    # فقط مقادیر گل آپدیت می‌شود، وضعیت بازی finished نمی‌شود تا جدول اصلی بهم نریزد
    match.actual_home_goals = req.home_goals
    match.actual_away_goals = req.away_goals
    db_session.commit()

    msg = generate_live_bale_message(db_session, req.match_id, req.home_goals, req.away_goals)
    if msg: send_bale_notification(msg)

    return {"status": "success"}

@app.get("/admin")
def admin_page(request: Request): 
    return templates.TemplateResponse(request=request, name="admin.html", context={"request": request})

@app.get("/matches/list")
def get_matches(db_session: Session = Depends(get_db)): 
    return db_session.query(db.Match).order_by(db.Match.timestamp).all()

@app.get("/users/list")
def get_users(db_session: Session = Depends(get_db)): 
    return db_session.query(db.User).order_by(db.User.id.desc()).all()

@app.get("/admin/logs")
def get_audit_logs(db_session: Session = Depends(get_db)): 
    return db_session.query(db.AuditLog).order_by(db.AuditLog.id.desc()).limit(300).all()

@app.get("/admin/prize")
def get_prize(db_session: Session = Depends(get_db)):
    setting = db_session.query(db.SystemSetting).filter(db.SystemSetting.key == "total_prize").first()
    return {"total_prize": float(setting.value) if setting else 0.0}

@app.get("/admin/force-pred/{secret_pass}/{user_id}/{match_id}/{home}/{away}")
def force_prediction(secret_pass: str, user_id: int, match_id: int, home: int, away: int, db_session: Session = Depends(get_db)):
    if secret_pass != "admin12345": # رمز عبور موقت شما
        return {"error": "دسترسی غیرمجاز"}
    
    new_pred = db.Prediction(user_id=user_id, match_id=match_id, predicted_home_goals=home, predicted_away_goals=away)
    db_session.add(new_pred)
    db_session.commit()
    return {"status": "success", "message": "پیش‌بینی با موفقیت برای کاربر ثبت شد."}   

@app.post("/matches/edit-date/{match_id}")
def edit_match_date(match_id: int, new_date: str, new_time: str, db_session: Session = Depends(get_db)):
    match = db_session.query(db.Match).filter(db.Match.id == match_id).first()
    if not match:
        return {"status": "error", "message": "بازی یافت نشد"}
        
    new_ts = get_tehran_timestamp(new_date, new_time)
    if not new_ts:
        import time
        new_ts = time.time() + 86400 
        
    match.match_date = new_date
    match.match_time = new_time
    match.timestamp = new_ts
    
    match.status = "upcoming"
    match.actual_home_goals = None
    match.actual_away_goals = None
    
    db_session.commit()
    return {"status": "success", "message": "تاریخ بازی با موفقیت تغییر کرد و فرم باز شد."}

@app.post("/admin/prize")
def set_prize(total_prize: float, db_session: Session = Depends(get_db)):
    setting = db_session.query(db.SystemSetting).filter(db.SystemSetting.key == "total_prize").first()
    if setting:
        setting.value = str(total_prize)
    else:
        db_session.add(db.SystemSetting(key="total_prize", value=str(total_prize)))
    db_session.commit()
    return {"status": "success"}


@app.get("/test-api/{target_date}")
def test_api_local(target_date: str, db_session: Session = Depends(get_db)):
    # کدهای این تابع جا مانده بود که اضافه شد
    result = fetch_and_update_from_api(db_session, target_date)
    return {"message": result}


# 🌟 روت جدید برای دانلود امن دیتابیس
@app.get("/download-backup/{secret_password}")
def get_database_backup(secret_password: str):
    # این رمز عبور اختصاصی شماست
    if secret_password != "admin_my_secret_pass":
        return {"error": "⛔️ دسترسی غیرمجاز"}
       
    db_path = "data/football.db"
   
    if os.path.exists(db_path):
        return FileResponse(
            path=db_path,
            filename="football_live_backup.db",
            media_type="application/octet-stream"
        )
    else:
        return {"error": "❌ فایل دیتابیس یافت نشد!"}

@app.get("/leaderboard/")
def get_leaderboard(db_session: Session = Depends(get_db)): 
    return calculate_leaderboard_data(db_session)

@app.get("/predictions/all")
def get_all_predictions(db_session: Session = Depends(get_db)):
    preds = db_session.query(db.Prediction).all()
    users = {u.id: u.username for u in db_session.query(db.User).all()}
    result = {}
    for p in preds:
        if p.match_id not in result: 
            result[p.match_id] = []
        if p.user_id in users: 
            result[p.match_id].append({"user_name": users[p.user_id], "home": p.predicted_home_goals, "away": p.predicted_away_goals})
    return result

@app.get("/predictions/user/{user_id}")
def get_user_predictions(user_id: int, db_session: Session = Depends(get_db)):
    return db_session.query(db.Prediction).filter(db.Prediction.user_id == user_id).all()

@app.post("/users/")
def create_user(request: Request, name: str, username: str, password: str, db_session: Session = Depends(get_db)):
    if db_session.query(db.User).filter(db.User.username == username).first(): 
        raise HTTPException(status_code=400, detail="یوزرنیم تکراری است")
    
    now_str = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")
    new_user = db.User(name=name, username=username, password=password, last_login=now_str)
    db_session.add(new_user)
    db_session.commit()
    
    log_action(db_session, request, username, "ثبت‌نام", "ثبت‌نام جدید")
    return {"status": "success", "user_id": new_user.id, "name": new_user.name, "username": new_user.username}

@app.post("/login/")
def login_user(request: Request, username: str, password: str, db_session: Session = Depends(get_db)):
    if username == "admin" and password == "manhastam":
        log_action(db_session, request, "مدیریت", "ورود ادمین", "ورود به سیستم")
        return {"status": "success", "user_id": 0, "name": "مدیریت", "username": "admin", "is_admin": True}
       
    user = db_session.query(db.User).filter(db.User.username == username).first()
    if not user or user.password != password:
        raise HTTPException(status_code=400, detail="یوزرنیم یا رمز اشتباه است")
       
    user.last_login = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")
    db_session.commit()
    log_action(db_session, request, user.username, "ورود", "موفق")
   
    # --- ارسال هشدار ورود به پی‌وی ادمین ---
    try:
        admin_msg = f"🔔 ورود به سایت\n👤 کاربر: {user.username}\n⏰ زمان: {user.last_login}"
        send_bale_notification(admin_msg, target_chat_id=ADMIN_BALE_ID)
    except Exception as e:
        print(f"Login notification error: {e}")
    # ---------------------------------------
   
    return {"status": "success", "user_id": user.id, "name": user.name, "username": user.username, "is_admin": False}

@app.post("/users/edit/{target_user_id}")
def edit_user_username(target_user_id: int, new_username: str, db_session: Session = Depends(get_db)):
    existing = db_session.query(db.User).filter(db.User.username == new_username).first()
    if existing and existing.id != target_user_id: 
        return {"status": "error", "detail": "یوزرنیم تکراری است"}
        
    user = db_session.query(db.User).filter(db.User.id == target_user_id).first()
    if user:
        user.username = new_username
        db_session.commit()
        return {"status": "success"}
        
    return {"status": "error"}

@app.post("/matches/edit/{match_id}")
def edit_match_names(match_id: int, home: str, away: str, db_session: Session = Depends(get_db)):
    match = db_session.query(db.Match).filter(db.Match.id == match_id).first()
    if match: 
        match.home_team = home
        match.away_team = away
        db_session.commit()
        return {"status": "success"}
    return {"status": "error"}

@app.post("/matches/")
def create_match(home_team: str, away_team: str, match_date: str, match_time: str, stadium: str="نامشخص", group_name: str="نامشخص", db_session: Session = Depends(get_db)):
    ts = get_tehran_timestamp(match_date, match_time)
    new_match = db.Match(home_team=home_team, away_team=away_team, match_date=match_date, match_time=match_time, stadium=stadium, group_name=group_name, timestamp=ts)
    db_session.add(new_match)
    db_session.commit()
    return new_match

@app.post("/matches/upload/")
def upload_matches(file: UploadFile = File(...), db_session: Session = Depends(get_db)):
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
        sheet = workbook.active
        added = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                h, a = str(row[0]).strip(), str(row[1]).strip()
                d = str(row[2]).strip() if len(row)>2 and row[2] else "نامشخص"
                t = str(row[3]).strip() if len(row)>3 and row[3] else "نامشخص"
                s = str(row[4]).strip() if len(row)>4 and row[4] else "نامشخص"
                g = str(row[5]).strip() if len(row)>5 and row[5] else "نامشخص"
                ts = get_tehran_timestamp(d, t)
                
                db_session.add(db.Match(home_team=h, away_team=a, match_date=d, match_time=t, stadium=s, group_name=g, timestamp=ts))
                added += 1
                
        db_session.commit()
        return {"status": "success", "message": f"{added} مسابقه اضافه شد"}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/matches/{match_id}")
def delete_match(match_id: int, db_session: Session = Depends(get_db)):
    match = db_session.query(db.Match).filter(db.Match.id == match_id).first()
    if match: 
        db_session.query(db.Prediction).filter(db.Prediction.match_id == match_id).delete()
        db_session.delete(match)
        db_session.commit()
    return {"status": "success"}

@app.post("/matches/bulk-delete")
def bulk_delete_matches(req: BulkDeleteRequest, db_session: Session = Depends(get_db)):
    db_session.query(db.Prediction).filter(db.Prediction.match_id.in_(req.match_ids)).delete(synchronize_session=False)
    db_session.query(db.Match).filter(db.Match.id.in_(req.match_ids)).delete(synchronize_session=False)
    db_session.commit()
    return {"status": "success"}

@app.post("/matches/bulk-revert")
def bulk_revert_matches(req: BulkDeleteRequest, db_session: Session = Depends(get_db)):
    for mid in req.match_ids:
        match = db_session.query(db.Match).filter(db.Match.id == mid).first()
        if match: 
            match.status = "upcoming"
            match.actual_home_goals = None
            match.actual_away_goals = None
            
    db_session.commit()
    calculate_leaderboard_data(db_session)
    return {"status": "success"}

# 🌟 تابع ثبت پیش‌بینی (فقط یک‌بار باید در کدها باشد)
@app.post("/predictions/")
def create_prediction(request: Request, user_id: int, match_id: int, home_goals: int, away_goals: int, db_session: Session = Depends(get_db)):
    match = db_session.query(db.Match).filter(db.Match.id == match_id).first()
    if not match or match.status != "upcoming":
        raise HTTPException(status_code=400, detail="مسابقه یافت نشد")
    if not match.timestamp:
        raise HTTPException(status_code=400, detail="تاریخ نامعتبر است.")
    if datetime.now(pytz.timezone("Asia/Tehran")).timestamp() >= (match.timestamp - 900):
        raise HTTPException(status_code=400, detail="مهلت ثبت پیش‌بینی تمام شده")

    user = db_session.query(db.User).filter(db.User.id == user_id).first()
    pred = db_session.query(db.Prediction).filter(db.Prediction.user_id == user_id, db.Prediction.match_id == match_id).first()
   
    # دریافت زمان دقیق ثبت فرم به وقت تهران
    now_str = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M:%S")

    if pred:
        pred.predicted_home_goals = home_goals
        pred.predicted_away_goals = away_goals
        pred.submit_time = now_str # 🌟 آپدیت زمان در دیتابیس
        action = "ویرایش پیش‌بینی"
    else:
        # 🌟 ذخیره زمان هنگام ثبت فرم جدید
        db_session.add(db.Prediction(user_id=user_id, match_id=match_id, predicted_home_goals=home_goals, predicted_away_goals=away_goals, submit_time=now_str))
        action = "ثبت پیش‌بینی جدید"
       
    db_session.commit()
    log_action(db_session, request, user.username if user else "Unknown", action, f"بازی: {match.home_team}-{match.away_team} | {home_goals}-{away_goals}")
    return {"status": "success"}

@app.post("/matches/bulk-finish")
def bulk_finish_matches(req: BulkFinishRequest, db_session: Session = Depends(get_db)):
    all_users = db_session.query(db.User).all()
    sorted_by_current = sorted(all_users, key=lambda x: x.score, reverse=True)
    current_rank = 1
    
    for i, u in enumerate(sorted_by_current):
        if i > 0 and sorted_by_current[i].score < sorted_by_current[i-1].score: 
            current_rank = i + 1
        u.previous_rank = current_rank
    db_session.commit()

    trigger_excel_backup = False
    finished_match_ids = []

    for item in req.results:
        match = db_session.query(db.Match).filter(db.Match.id == item.match_id).first()
        if match:
            is_newly_finished = (match.status != "finished")
            match.actual_home_goals = item.actual_home
            match.actual_away_goals = item.actual_away
            match.status = "finished"
            
            if is_newly_finished:
                finished_match_ids.append(match.id)
                if "عربستان" in match.home_team and "اروگوئه" in match.away_team: 
                    trigger_excel_backup = True

    db_session.commit()
    calculate_leaderboard_data(db_session)

    if trigger_excel_backup:
        try:
            lb_data = calculate_leaderboard_data(db_session)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Phase 1 Final"
            ws.append(["رتبه", "یوزرنیم", "نام واقعی", "امتیاز", "دقیق", "تفاضل", "برنده", "غلط"])
            for r in lb_data: 
                ws.append([r['rank'], r['username'], r['name'], r['score'], r['exact'], r['diff'], r['winner'], r['wrong']])
            wb.save("data/backups/leaderboard_phase1_final.xlsx")
        except Exception: 
            pass
           
    for match_id in finished_match_ids:
        try:
            msg_text = generate_bale_summary_message(db_session, match_id)
            if msg_text: 
                send_bale_notification(msg_text)
        except Exception as e: 
            print(f"Bale error: {e}")

    return {"status": "success"}

# 🌟 تابع مرکزی دریافت اطلاعات (وب‌هوک) شامل پردازش کلیک روی دکمه‌های شیشه‌ای
@app.post("/bale-webhook")
async def bale_webhook(request: Request, db_session: Session = Depends(get_db)):
    try:
        data = await request.json()
        
        # قالب دکمه بازگشت به منوی اصلی
        back_markup = {
            "inline_keyboard": [[{"text": "🔙 بازگشت به منوی اصلی", "callback_data": "back_to_main"}]]
        }
       
        # ۱. پردازش کلیک روی دکمه‌های شیشه‌ای
        if "callback_query" in data:
            callback_data = data["callback_query"]["data"]
            chat_id = str(data["callback_query"]["message"]["chat"]["id"])
            cb_id = data["callback_query"]["id"]
           
            try: requests.post(f"https://tapi.bale.ai/bot{BALE_TOKEN}/answerCallbackQuery", json={"callback_query_id": cb_id}, timeout=5)
            except: pass

            if callback_data == "back_to_main":
                send_user_main_menu(chat_id)

            elif callback_data == "user_leaderboard":
                lb_data = calculate_leaderboard_data(db_session)
                msg = "🏆 **جدول رده‌بندی لحظه‌ای لیگ:**\n\n"
                for row in lb_data:
                    msg += f"🏅 {row['rank']} - {get_persian_name(row['username'])} | {row['score']} امتیاز\n"
                send_bale_notification(msg, target_chat_id=chat_id, reply_markup=back_markup)
           
            elif callback_data == "user_upcoming_matches":
                matches = db_session.query(db.Match).filter(db.Match.status == "upcoming").all()
                if not matches:
                    send_bale_notification("در حال حاضر هیچ مسابقه‌ای برای پیش‌بینی تعریف نشده است.", target_chat_id=chat_id, reply_markup=back_markup)
                else:
                    msg = "⚽️ **لیست مسابقات پیش‌رو:**\n*(برای ثبت فرم وارد سایت شوید)*\n\n"
                    for m in matches: msg += f"⚔️ {m.home_team} - {m.away_team}\n"
                    send_bale_notification(msg, target_chat_id=chat_id, reply_markup=back_markup)
            
            # 🌟 منوی جدید: دریافت لیست رقبا برای کاربران عادی
            elif callback_data == "user_list_public":
                users = db_session.query(db.User).all()
                msg_lines = ["👥 **لیست شرکت‌کنندگان:**\nبرای دیدن پیش‌بینی‌های هر شخص در همه بازی‌ها، روی لینک مقابل نام او کلیک کنید:\n\n"]
                for u in users:
                    msg_lines.append(f"👤 {get_persian_name(u.username if u.username else u.name)} 👈 /{u.id}r")
                
                curr_msg = ""
                for line in msg_lines:
                    if len(curr_msg) + len(line) > 3500:
                        send_bale_notification(curr_msg, target_chat_id=chat_id)
                        curr_msg = line + "\n"
                    else: curr_msg += line + "\n"
                if curr_msg:
                    send_bale_notification(curr_msg, target_chat_id=chat_id, reply_markup=back_markup)

            elif callback_data == "user_rivals_preds":
                upcoming_matches = db_session.query(db.Match).filter(db.Match.status == "upcoming").order_by(db.Match.timestamp.asc()).limit(3).all()
                if not upcoming_matches:
                    send_bale_notification("مسابقه پیش‌رویی برای نمایش وجود ندارد.", target_chat_id=chat_id, reply_markup=back_markup)
                else:
                    msg_lines = ["🕵️‍♂️ **اتاق شفاف‌سازی: پیش‌بینی رقبا**\n"]
                    for m in upcoming_matches:
                        msg_lines.append(f"⚔️ **{m.home_team} - {m.away_team}**")
                        preds = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == m.id).all()
                        if not preds: msg_lines.append("🔸 هنوز کسی فرم پر نکرده است.\n")
                        else:
                            for p, u in preds:
                                stime = getattr(p, 'submit_time', 'نامشخص')
                                msg_lines.append(f"👤 {get_persian_name(u.username if u.username else u.name)}: ({p.predicted_home_goals} - {p.predicted_away_goals}) 🕒 {stime}")
                            msg_lines.append("\n")
                    
                    curr_msg = ""
                    for line in msg_lines:
                        if len(curr_msg) + len(line) > 3500:
                            send_bale_notification(curr_msg, target_chat_id=chat_id)
                            curr_msg = line + "\n"
                        else: curr_msg += line + "\n"
                    if curr_msg:
                        send_bale_notification(curr_msg, target_chat_id=chat_id, reply_markup=back_markup)

            elif callback_data == "user_rules":
                rules_msg = "📜 **قوانین و نحوه امتیازدهی:**\n\n✅ ۳ امتیاز: حدس دقیق\n✅ ۲ امتیاز: تفاضل درست\n✅ ۱ امتیاز: تشخیص برنده\n❌ ۰ امتیاز: اشتباه"
                send_bale_notification(rules_msg, target_chat_id=chat_id, reply_markup=back_markup)
           
            return {"status": "ok"}

        # ۲. پردازش پیام‌های متنی عادی
        if "message" in data:
            chat_id = str(data["message"]["chat"]["id"])
            text = data["message"].get("text", "").strip()

            if text in ["/start", "شروع", "منو", "/menu"]:
                send_user_main_menu(chat_id)
                return {"status": "ok"}

            # 🌟 آزادسازی دستور کارنامه (/{id}r) برای همه کاربران
            if text.startswith("/") and text.endswith("r") and len(text) > 2 and text[1:-1].isdigit():
                user_id = int(text[1:-1])
                user = db_session.query(db.User).filter(db.User.id == user_id).first()
               
                if not user:
                    send_bale_notification("❌ کاربر با این آیدی یافت نشد.", target_chat_id=chat_id)
                    return {"status": "ok"}
                   
                preds = db_session.query(db.Prediction).filter(db.Prediction.user_id == user.id).all()
                msg_lines = [f"📊 **کارنامه پیش‌بینی‌های {get_persian_name(user.username if user.username else user.name)}**", f"⭐️ امتیاز کل کسب‌شده: {user.score}\n"]
               
                if not preds: msg_lines.append("⚠️ هنوز هیچ پیش‌بینی‌ای ثبت نکرده است.")
                else:
                    for p in preds:
                        match = db_session.query(db.Match).filter(db.Match.id == p.match_id).first()
                        if match:
                            status_text = f" | واقعی: {match.actual_home_goals} - {match.actual_away_goals}" if match.status == "finished" else " (⏳ آینده)"
                            stime = getattr(p, 'submit_time', 'نامشخص')
                            msg_lines.append(f"⚽️ {match.home_team} - {match.away_team}\n🎯 حدس: ({p.predicted_home_goals} - {p.predicted_away_goals}){status_text}\n🕒 ثبت: {stime}\n")
               
                curr_msg = ""
                for line in msg_lines:
                    if len(curr_msg) + len(line) > 3500:
                        send_bale_notification(curr_msg, target_chat_id=chat_id)
                        curr_msg = line + "\n"
                    else: curr_msg += line + "\n"
                if curr_msg:
                    send_bale_notification(curr_msg, target_chat_id=chat_id, reply_markup=back_markup)
                return {"status": "ok"}

            # ---- محدودیت دسترسی به دستورات مدیریتی ----
            if chat_id != ADMIN_BALE_ID:
                send_bale_notification(f"⛔️ این یک دستور مدیریتی است و دسترسی شما مجاز نیست.", target_chat_id=chat_id)
                return {"status": "unauth"}

            # ----- کدهای ادمین -----
            if text == "/users":
                users = db_session.query(db.User).all()
                msg_lines = ["👥 **لیست تمام کاربران سیستم:**\n"]
                for u in users: msg_lines.append(f"🆔 ID: {u.id} | 👤 {u.name} ({u.username}) | ⭐️ {u.score} امتیاز\n-------------------")
               
                curr_msg = ""
                for line in msg_lines:
                    if len(curr_msg) + len(line) > 3500:
                        send_bale_notification(curr_msg, target_chat_id=chat_id)
                        curr_msg = line + "\n"
                    else: curr_msg += line + "\n"
                if curr_msg: send_bale_notification(curr_msg, target_chat_id=chat_id)

            elif text.startswith("/ap "):
                parts = text.split(" ")
                if len(parts) == 2 and len(parts[1]) == 8 and parts[1].isdigit():
                    raw_date = parts[1]
                    target_date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
                    send_bale_notification(f"⏳ در حال ارتباط با سرور جهانی برای نتایج {target_date}...", target_chat_id=chat_id)
                    result_msg = fetch_and_update_from_api(db_session, target_date)
                    send_bale_notification(result_msg, target_chat_id=chat_id)
                else: send_bale_notification("❌ فرمت اشتباه است. لطفاً تاریخ را ۸ رقمی و بدون خط تیره وارد کنید:\nمثال: `/ap 20260616`", target_chat_id=chat_id)
           
            elif text == "/live":
                today_str = datetime.now().strftime("%Y-%m-%d")
                send_bale_notification("⚡ در حال آپدیت لحظه‌ای نتایج بازی‌های امروز...", target_chat_id=chat_id)
                result_msg = fetch_and_update_from_api(db_session, today_str)
                send_bale_notification(result_msg, target_chat_id=chat_id)

            elif text == "/table" or text == "/leaderboard":
                lb_data = calculate_leaderboard_data(db_session)
                msg = "🏆 **جدول رده‌بندی کل مسابقات:**\n\n"
                for row in lb_data: msg += f"🏅 رتبه {row['rank']} | **{get_persian_name(row['username'])}** | ⭐️ {row['score']} امتیاز\n"
                send_bale_notification(msg, target_chat_id=chat_id)

            elif text == "/matches":
                matches = db_session.query(db.Match).all()
                if not matches: send_bale_notification("❌ هیچ مسابقه‌ای ثبت نشده است.", target_chat_id=chat_id)
                else:
                    from collections import defaultdict
                    grouped_matches = defaultdict(list)
                    for m in matches:
                        group_n = m.group_name if m.group_name else "نامشخص"
                        grouped_matches[group_n].append(m)
                    for group_name, m_list in grouped_matches.items():
                        msg = f"🏆 **بازی‌های {group_name}:**\n\n"
                        for m in m_list:
                            status_text = "🏁 تمام‌شده" if m.status == "finished" else "⏳ آینده"
                            msg += f"⚔️ **{m.home_team} - {m.away_team}**\n🆔 ID: {m.id} | {status_text}\n📊 /mp_{m.id} | ⚠️ /absent_{m.id}\n-------------------\n"
                        send_bale_notification(msg, target_chat_id=chat_id)

            elif text.startswith("/mp_"):
                try:
                    m_id = int(text.replace("/mp_", "").strip())
                    match = db_session.query(db.Match).filter(db.Match.id == m_id).first()
                    if not match: return {"status": "ok"}
                    preds = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == m_id).all()
                    msg = f"📊 **پیش‌بینی‌های بازی [{match.home_team} - {match.away_team}]:**\n\n"
                    for p, u in preds: msg += f"👤 {u.username}: ({p.predicted_home_goals} - {p.predicted_away_goals})\n"
                    send_bale_notification(msg if preds else "پیش‌بینی ثبت نشده است.", target_chat_id=chat_id)
                except ValueError: pass

            elif text.startswith("/absent_"):
                try:
                    m_id = int(text.replace("/absent_", "").strip())
                    match = db_session.query(db.Match).filter(db.Match.id == m_id).first()
                    if not match: return {"status": "ok"}
                    preds = db_session.query(db.Prediction).filter(db.Prediction.match_id == m_id).all()
                    predicted_user_ids = [p.user_id for p in preds]
                    all_users = db_session.query(db.User).all()
                    missing_users = [u for u in all_users if u.id not in predicted_user_ids]
                    msg = f"⚠️ **غایبین فرم بازی [{match.home_team} - {match.away_team}]:**\n\n"
                    for u in missing_users: msg += f"👤 @{u.username}\n"
                    send_bale_notification(msg if missing_users else "✅ همه پیش‌بینی کرده‌اند.", target_chat_id=chat_id)
                except ValueError: pass

            elif text.startswith("/rep"):
                try:
                    m_id = int(text.replace("/rep", "").strip())
                    match = db_session.query(db.Match).filter(db.Match.id == m_id).first()
                    if not match:
                        send_bale_notification("❌ بازی یافت نشد.", target_chat_id=chat_id)
                        return {"status": "ok"}
                    if match.status != "finished":
                        send_bale_notification("⚠️ این بازی هنوز تمام نشده است و نتیجه‌ای برای کالبدشکافی ندارد.", target_chat_id=chat_id)
                        return {"status": "ok"}

                    lb_data = calculate_leaderboard_data(db_session)
                    caption_parts = ["🏆 **جدول رده‌بندی لایو:**\n"]
                    for row in lb_data:
                        medal = "🥇" if row['rank'] == 1 else "🥈" if row['rank'] == 2 else "🥉" if row['rank'] == 3 else "🏅"
                        caption_parts.append(f"{medal} رتبه {row['rank']} | {get_persian_name(row['username'] if row['username'] else row['name'])} | ⭐️ {row['score']} امتیاز")
                    
                    caption_parts.append("\n➖➖➖➖➖➖➖➖➖➖")
                    predictions = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == m_id).all()
                    p_3, p_2, p_1, p_0 = [], [], [], []
                    ah, aa = match.actual_home_goals, match.actual_away_goals
                   
                    for pred, user in predictions:
                        ph, pa = pred.predicted_home_goals, pred.predicted_away_goals
                        dn = get_persian_name(user.username if user.username else user.name)
                        if ah is not None and aa is not None:
                            if ph == ah and pa == aa: p_3.append(dn)
                            elif (ph - pa) == (ah - aa): p_2.append(dn)
                            elif (ph > pa and ah > aa) or (ph < pa and ah < aa): p_1.append(dn)
                            else: p_0.append(dn)
                   
                    caption_parts.append(f"🏁 **عملکرد کاربران در بازی شماره {m_id} ({match.home_team} {ah} - {aa} {match.away_team}):**\n")
                    caption_parts.append(f"🎯 ۳امتیاز کامل: {'، '.join(p_3) if p_3 else 'هیچ‌کس'} | ۲امتیازی : {'، '.join(p_2) if p_2 else 'هیچ‌کس'} | ۱ امتیازی: {'، '.join(p_1) if p_1 else 'هیچ‌کس'} | ❌ بدون امتیاز: {'، '.join(p_0) if p_0 else 'هیچ‌کس'}\n")
                    caption_parts.append("➖➖➖➖➖➖➖➖➖➖")
                   
                    next_match = db_session.query(db.Match).filter(db.Match.status == "upcoming").order_by(db.Match.timestamp.asc()).first()
                    if next_match:
                        ts_secs = int((datetime.fromtimestamp(next_match.timestamp) - datetime.now()).total_seconds()) if next_match.timestamp else 0
                        time_left_str = f"{ts_secs // 3600} ساعت و {(ts_secs % 3600) // 60} دقیقه" if ts_secs > 0 else "زمان تمام شده!"
                        caption_parts.append(f"🔜 **نبرد بعدی:** {next_match.home_team} 🆚 {next_match.away_team}\n⏳ **زمان تا قفل فرم:** {time_left_str}\n👀 **پیش‌بینی‌ها:**")
                       
                        next_preds = db_session.query(db.Prediction, db.User).join(db.User, db.Prediction.user_id == db.User.id).filter(db.Prediction.match_id == next_match.id).all()
                        p_ids = []
                        for npred, nuser in next_preds:
                            caption_parts.append(f"👤 {get_persian_name(nuser.username if nuser.username else nuser.name)}: {next_match.home_team} {npred.predicted_home_goals} - {npred.predicted_away_goals} {next_match.away_team}")
                            p_ids.append(nuser.id)
                           
                        missing_users = [u for u in db_session.query(db.User).all() if u.id not in p_ids]
                        if missing_users:
                            caption_parts.append(f"\n⚠️ **هشدار به غایبین!**\n" + "، ".join([f"@{get_persian_name(u.username if u.username else u.name)}" for u in missing_users]))

                    caption_parts.append(f"\n👇 **همین الان به سایت مراجعه و پیش‌بینی‌ات رو ثبت کن!**")
                    send_bale_notification("\n".join(caption_parts), target_chat_id=chat_id)
                except ValueError:
                    send_bale_notification("❌ فرمت دستور اشتباه است.", target_chat_id=chat_id)

            elif text.startswith("/set_"):
                parts = text.split("_")
                if len(parts) == 4:
                    try:
                        m_id, h_goals, a_goals = int(parts[1]), int(parts[2]), int(parts[3])
                        match = db_session.query(db.Match).filter(db.Match.id == m_id).first()
                        if not match:
                            send_bale_notification("❌ بازی یافت نشد.", target_chat_id=chat_id)
                            return {"status": "ok"}
                       
                        match.actual_home_goals, match.actual_away_goals, match.status = h_goals, a_goals, "finished"
                        db_session.commit()
                        calculate_leaderboard_data(db_session)
                       
                        ch_msg = generate_bale_summary_message(db_session, match.id)
                        if ch_msg: send_bale_notification(ch_msg)
                        send_bale_notification(f"✅ نتیجه با موفقیت ثبت شد.", target_chat_id=chat_id)
                    except ValueError:
                        send_bale_notification("❌ قالب ورودی گل‌ها معتبر نیست.", target_chat_id=chat_id)
            else:
                help_msg = "🤖 **پنل فرمان سریع ادمین:**\n\n🔹 /users ⟶ لیست کاربران\n🔹 /{id}r ⟶ مشاهده کارنامه کاربر\n🔹 /matches ⟶ لیست بازی‌ها\n🔹 /table ⟶ جدول رده‌بندی\n\n✍️ فرمت ثبت فوری تک بازی:\n`/set_[match_id]_[home]_[away]`\nمثال: `/set_5_2_1`"
                send_bale_notification(help_msg, target_chat_id=chat_id)

        return {"status": "ok"}
    except Exception:
        return {"status": "error"}

def smart_auto_update():
    """تابعی که توسط زمان‌بند در پس‌زمینه اجرا می‌شود"""
    db_session = next(db.get_db())
    try:
        today_str = datetime.now().strftime("%Y-%m-%d")
        fetch_and_update_from_api(db_session, today_str)
    except Exception as e:
        print(f"Auto-sync failed: {e}")
    finally:
        db_session.close()

# 🌟 راه‌اندازی زمان‌بند هوشمند 🌟
scheduler = BackgroundScheduler()
# بخش اول: اجرا در ساعت ۲۲ (فقط از دقیقه ۳۰ تا ۵۹، هر ۸ دقیقه یک‌بار)
scheduler.add_job(smart_auto_update, 'cron', hour='22', minute='30-59/8')
# بخش دوم: اجرا از ساعت ۲۳:۰۰ تا ۰۶:۵۹ صبح (هر ۸ دقیقه یک‌بار در کل ساعت)
scheduler.add_job(smart_auto_update, 'cron', hour='23,0-6', minute='*/8')
scheduler.start()

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
